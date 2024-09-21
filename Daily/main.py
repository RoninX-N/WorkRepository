# -*- coding: utf-8 -*-
# @Time : 2023/12/6 13:46
# @Author : RoninX

import pandas as pd
from openpyxl import load_workbook
import numpy as np
import cems
import inorganization
from datetime import datetime,timedelta
import re

# <editor-fold desc='报警表常量'>
EVENTTYPE = "类型"
EQUIPMENTNAME = "设备名称"
EQUIPMENTTYPE = "设备类型"
FACTORY = "所属分厂"
PRODUCTIONPROCESS = "生产工序"
ALARMTYPE = "报警类型"
ALARMSTATUS = "报警状态"
ALARMDETAILS = "报警详情"
STARTTIME = "起始时间"
ALARMCAUSE = "建议排查原因"
ALARMTIME = "报警时间"
MONITOREDVALUE = "监测值"
LIMITS = "限值"
FACTORS = "污染因子"
SO2 = "二氧化硫"
NOX = "氮氧化物"
PARTICAL = "烟尘"
# </editor-fold>

# <editor-fold desc='各厂部名称定义'>
IRONMAKE = "炼铁厂"
STEELMAKE = "炼钢厂"
STRIPS = "条材厂"
HOTROLL = "热轧厂"
COLDROLL = "冷轧厂"
SILICONSTEEL = "硅钢部"
STEELELECTRICITY = "钢电公司"
TRANSPORT = "运输部"
EAEP = "能环部"
WINSTEEL = "Winsteel"
# </editor-fold>

# <editor-fold desc='事件类型与设备类型定义'>
CALIBRATION = "检校"
STOPPAGE = "启停"
OPERATECONDITION = "工况"
CEMSFAULT = "CEMS故障"
PROCESSED = "已处理"
NOTPROCESSED = "未处理"
APPROVED = "审核通过"
GOVERNANCEFACILITIES = "治理设施"
TVOC = "TVOC监测仪"
UNLINKAGE = "未联动开启"
DIFFERENTIALPRESSUREOVERRUN = "除尘器连续超出额定压差范围"
FANSCURRENTOVERRUN = "风机电流连续超过额定电流"
# </editor-fold>

# <editor-fold desc='文件路径定义'>
ONLINEFILEPATH = "D:\\elseFile\\work\\Codeworkbench\\ChartTemplate\\报警历史（到2024.9.20下午4点）.xlsx"
SmartandenvironmentallyfriendlyINPUTFILEPATH = "D:\\elseFile\\work\\Codeworkbench\\ChartTemplate\\（模板）智慧环保系统运行日报.xlsx"
SmartandenvironmentallyfriendlyOUTPUTFILEPATH = "D:\\elseFile\\work\\Codeworkbench\\ChartOutPut\\（模板）智慧环保系统运行日报.xlsx"
calibrationplanFILEPATH = "D:\\elseFile\\work\\Codeworkbench\\ChartTemplate\\校准计划表(10月更新).xlsx"
# </editor-fold>

def ppt_summary(cemsarray,Smartandenvironmentallyfriendlysheet0,hourly_average_position,abnormal_cause_positon):
    # <editor-fold desc='定义PPT的两个字符串跟各种数组、临时变量、正则表达式'>
    hourly_average_string = "累计超标"
    abnormal_analysis_string = ""

    PPTarray = [[EVENTTYPE,EQUIPMENTNAME,ALARMTIME,FACTORS,MONITOREDVALUE,LIMITS,ALARMCAUSE]]
    temp_time_string = ""
    temp_monitoredvalue_string = ""
    temp_limits_string = ""
    pattern0 = re.compile(r'\(([^)]*)\)')
    pattern1 = re.compile(r'\)(.*?)超')
    pattern2 = re.compile(r"监测值\(([\d.]+)\)")
    pattern3 = re.compile(r"限值\(([\d.]+)\)")
    # </editor-fold>
    for row in cemsarray[1:]:
        newrow = [row[0],row[1],pattern1.findall(row[4])[0],pattern0.search(row[4]).group(1),pattern2.search(row[4]).group(1),pattern3.search(row[4]).group(1),row[5]]
        PPTarray = np.vstack((PPTarray,newrow))
    # <editor-fold desc='定义某个时间点，后面筛选在这个时间点之后的数据'> 这里注释里改时间代码
    current_datetime = datetime.now()
    start_time = datetime(current_datetime.year,current_datetime.month,current_datetime.day,6,0)
    # 如果是非当天做的日报，就用下面这一行修改日期
    # start_time = datetime(2024, 9, 1, 6, 0)
    if start_time.weekday() == 6:
        start_time = start_time - timedelta(days=2)
    PPTarray_filtered = []
    for row in PPTarray[1:]:
        date_str = row[2]
        date = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
        if date >= start_time:
            PPTarray_filtered.append(row)
    # </editor-fold>

    # <editor-fold desc='定义计数变量与各条信息数组'>
    allexceeded = 0
    calibration = 0
    stoppage = 0
    operatecondition = 0
    abnormal = 0
    caarray = [[EQUIPMENTNAME,FACTORS]]
    uniqueca = []
    starray = [[EQUIPMENTNAME,FACTORS]]
    uniquest = []
    oparray = [[EQUIPMENTNAME,ALARMTIME,FACTORS,MONITOREDVALUE,LIMITS,ALARMCAUSE]]
    uniqueop = []
    abarray = [[EQUIPMENTNAME,ALARMTIME,FACTORS,MONITOREDVALUE,LIMITS,ALARMCAUSE]]
    uniqueab = []
    # </editor-fold>

    # <editor-fold desc='将小时均值异常情况信息整理并填入相应单元格'>
    # 这个for循环整理各种超标类型的数组
    for row in PPTarray_filtered:
        if datetime.strptime(row[2],"%Y-%m-%d %H:%M")>=start_time:
            if row[0] == CALIBRATION:
                allexceeded += 1
                calibration += 1
                newrow = [row[1],row[3]]
                caarray = np.vstack((caarray,newrow))
                uniqueca.append(row[1])
            elif row[0] == STOPPAGE:
                allexceeded += 1
                stoppage += 1
                newrow = [row[1],row[3]]
                starray = np.vstack((starray,newrow))
                uniquest.append(row[1])
            elif row[0] == OPERATECONDITION:
                allexceeded += 1
                operatecondition += 1
                newrowplus = [row[1],row[2],row[3],row[4],row[5],row[6]]
                oparray = np.vstack((oparray,newrowplus))
                uniqueop.append(row[1])
            elif row[0] == CEMSFAULT:
                allexceeded += 1
                abnormal += 1
                newrowplus = [row[1],row[2],row[3],row[4],row[5],row[6]]
                abarray = np.vstack((abarray,newrowplus))
                uniqueab.append(row[1])
    uniqueca = list(set(uniqueca))
    uniquest = list(set(uniquest))
    uniqueop = list(set(uniqueop))
    uniqueab = list(set(uniqueab))
    hourly_average_string += str(allexceeded) + "次，其中"
    if calibration != 0:
        hourly_average_string += "检校超标" + str(calibration) + "次，"
        for x in uniqueca:
            temp_factors_array = []
            temp_partical = 0
            temp_SO2 = 0
            temp_NOx = 0
            for y in PPTarray_filtered:
                if y[1] == x:
                    temp_factors_array.append(y[3])
                    if y[3] == PARTICAL:
                        temp_partical += 1
                    elif y[3] == SO2:
                        temp_SO2 += 1
                    elif y[3] == NOX:
                        temp_NOx += 1
            temp_factors_array = list(set(temp_factors_array))
            hourly_average_string += x
            if temp_partical != 0:
                hourly_average_string += "颗粒物" + str(temp_partical) + "次、"
            if temp_SO2 != 0:
                hourly_average_string += "二氧化硫" + str(temp_SO2) + "次、"
            if temp_NOx != 0:
                hourly_average_string += "氮氧化物" + str(temp_NOx) + "次、"
            hourly_average_string = hourly_average_string[:-1]
        hourly_average_string += "；"
    if stoppage != 0:
        hourly_average_string += "启停炉（机）超标" + str(stoppage) + "次，"
        for x in uniquest:
            temp_factors_array = []
            temp_partical = 0
            temp_SO2 = 0
            temp_NOx = 0
            for y in PPTarray_filtered:
                if y[1] == x:
                    temp_factors_array.append(y[3])
                    if y[3] == PARTICAL:
                        temp_partical += 1
                    elif y[3] == SO2:
                        temp_SO2 += 1
                    elif y[3] == NOX:
                        temp_NOx += 1
            temp_factors_array = list(set(temp_factors_array))
            hourly_average_string += x
            if temp_partical != 0:
                hourly_average_string += "颗粒物" + str(temp_partical) + "次、"
            if temp_SO2 != 0:
                hourly_average_string += "二氧化硫" + str(temp_SO2) + "次、"
            if temp_NOx != 0:
                hourly_average_string += "氮氧化物" + str(temp_NOx) + "次、"
            hourly_average_string = hourly_average_string[:-1]
        hourly_average_string += "；"
    if operatecondition != 0:
        hourly_average_string += "工况超标" + str(operatecondition) + "次，"
        for x in uniqueop:
            temp_factors_array = []
            temp_partical = 0
            temp_SO2 = 0
            temp_NOx = 0
            for y in PPTarray_filtered:
                if y[1] == x:
                    temp_factors_array.append(y[3])
                    if y[3] == PARTICAL:
                        temp_partical += 1
                    elif y[3] == SO2:
                        temp_SO2 += 1
                    elif y[3] == NOX:
                        temp_NOx += 1
            temp_factors_array = list(set(temp_factors_array))
            hourly_average_string += x
            if temp_partical != 0:
                hourly_average_string += "颗粒物" + str(temp_partical) + "次、"
            if temp_SO2 != 0:
                hourly_average_string += "二氧化硫" + str(temp_SO2) + "次、"
            if temp_NOx != 0:
                hourly_average_string += "氮氧化物" + str(temp_NOx) + "次、"
            hourly_average_string = hourly_average_string[:-1]
        hourly_average_string += "；"
    if abnormal != 0:
        hourly_average_string += "环保设备异常超标" + str(abnormal) + "次，"
        for x in uniqueab:
            temp_factors_array = []
            temp_partical = 0
            temp_SO2 = 0
            temp_NOx = 0
            for y in PPTarray_filtered:
                if y[1] == x:
                    temp_factors_array.append(y[3])
                    if y[3] == PARTICAL:
                        temp_partical += 1
                    elif y[3] == SO2:
                        temp_SO2 += 1
                    elif y[3] == NOX:
                        temp_NOx += 1
            temp_factors_array = list(set(temp_factors_array))
            hourly_average_string += x
            if temp_partical != 0:
                hourly_average_string += "颗粒物" + str(temp_partical) + "次、"
            if temp_SO2 != 0:
                hourly_average_string += "二氧化硫" + str(temp_SO2) + "次、"
            if temp_NOx != 0:
                hourly_average_string += "氮氧化物" + str(temp_NOx) + "次、"
            hourly_average_string = hourly_average_string[:-1]
        hourly_average_string += "；"
    hourly_average_string = hourly_average_string[:-1] + "。"
    print(hourly_average_string)
    Smartandenvironmentallyfriendlysheet0[hourly_average_position] = hourly_average_string
    # </editor-fold>
    # print(oparray)
    # print(abarray)
    if len(oparray) > 1 or len(abarray) > 1:
        cause = ""
        if len(oparray) > 1:
            for x in uniqueop:
                temp_value = []
                temp_limit = 0
                flag = 0
                for y in oparray[1:]:
                    if y[0] == x and y[2] == "烟尘":
                        flag = 1
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "颗粒物超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" +"\n"
                    flag = 0
                temp_value = []
                for y in oparray[1:]:
                    if y[0] == x and y[2] == "二氧化硫":
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "二氧化硫超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" + "\n"
                    flag = 0
                temp_value = []
                for y in oparray[1:]:
                    if y[0] == x and y[2] == "氮氧化物":
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "氮氧化物超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" + "\n"
                    flag = 0

        if len(abarray) > 1:
            for x in uniqueab:
                temp_value = []
                temp_limit = 0
                flag = 0
                for y in abarray[1:]:
                    if y[0] == x and y[2] == "烟尘":
                        flag = 1
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "颗粒物超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" +"\n"
                    flag = 0
                temp_value = []
                for y in abarray[1:]:
                    if y[0] == x and y[2] == "二氧化硫":
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "二氧化硫超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" + "\n"
                    flag = 0
                temp_value = []
                for y in abarray[1:]:
                    if y[0] == x and y[2] == "氮氧化物":
                        abnormal_analysis_string += datetime.strptime(y[1],"%Y-%m-%d %H:%M").strftime("%m月%d日 %H:%M") + "、"
                        cause = y[5]
                        temp_value.append(y[3])
                        temp_limit = y[4]
                if flag == 1:
                    abnormal_analysis_string = abnormal_analysis_string[:-1] + "，" + x + "氮氧化物超标，监测值"
                    for z in temp_value:
                        abnormal_analysis_string += str(z) + "、"
                    abnormal_analysis_string = abnormal_analysis_string[:-1]
                    abnormal_analysis_string += "，限值" + str(temp_limit) + "，原因为" + y[5] + "。" + "\n"
                    flag = 0
    else:
        abnormal_analysis_string = "无"
    Smartandenvironmentallyfriendlysheet0[abnormal_cause_positon] = abnormal_analysis_string

    print(abnormal_analysis_string)


def cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,factory,calibrationfactoryA,calibrationalarmC,calibrationdisposedD,calibrationundisposedE,calibrationalarmdetailF,
                        stoppagefactoryA,stoppagealarmC,stoppagedisposedD,stoppageundisposedE,stoppagealarmdetailF,
                        operateconditionfactoryA,operateconditionalarmC,operateconditiondisposedD,operateconditionundisposedE,operateconditionalarmdetailF,
                        cemsfaultfactoryA,cemsfaultalarmC,cemsfaultdisposedD,cemsfaultundisposedE,cemsfaultalarmdetailF):
    cems_calibration = 0
    cems_stoppage = 0
    cems_operatecondition = 0
    cems_cemsfault = 0
    for row in cemsarray:
        if (row[0] == CALIBRATION and row[2] == factory):
            cems_calibration = cems_calibration + 1
        elif (row[0] == STOPPAGE and row[2] == factory):
            cems_stoppage = cems_stoppage + 1
        elif (row[0] == OPERATECONDITION and row[2] == factory):
            cems_operatecondition = cems_operatecondition + 1
        elif (row[0] == CEMSFAULT and row[2] == factory):
            cems_cemsfault = cems_cemsfault + 1
    cems_instance = cems.Cems()
    # 分厂检校类报警详情处理
    if (cems_calibration != 0):
        cems_calibration_string = cems_instance.calibration(cemsarray, factory)
        # 将返回的字符串输入到日报表对应位置
        Smartandenvironmentallyfriendlysheet0[calibrationfactoryA] = factory
        Smartandenvironmentallyfriendlysheet0[calibrationalarmdetailF] = cems_calibration_string
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for alarmstatuscount in cemsarray:
            if alarmstatuscount[2] == factory and (alarmstatuscount[3] == PROCESSED or alarmstatuscount[3] == APPROVED) and alarmstatuscount[0] == CALIBRATION:
                disposed += 1
                alarmcount += 1
            if alarmstatuscount[2] == factory and alarmstatuscount[3] == NOTPROCESSED and alarmstatuscount[0] == CALIBRATION:
                undisposed += 1
                alarmcount += 1
        Smartandenvironmentallyfriendlysheet0[calibrationalarmC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[calibrationdisposedD] = disposed
        Smartandenvironmentallyfriendlysheet0[calibrationundisposedE] = undisposed

    # 这里做输出excel文件测试，需要注释掉
    # Smartandenvironmentallyfriendly_input_workbook.save(SmartandenvironmentallyfriendlyOUTPUTFILEPATH)
    # 分厂启停类报警详情处理
    if (cems_stoppage != 0):
        cems_stoppage_string = cems_instance.stoppage(cemsarray, factory)
        # 将返回的字符串输入到日报表对应位置
        Smartandenvironmentallyfriendlysheet0[stoppagefactoryA] = factory
        Smartandenvironmentallyfriendlysheet0[stoppagealarmdetailF] = cems_stoppage_string
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for alarmstatuscount in cemsarray:
            if alarmstatuscount[2] == factory and (alarmstatuscount[3] == PROCESSED or alarmstatuscount[3] == APPROVED) and alarmstatuscount[0] == STOPPAGE:
                disposed += 1
                alarmcount += 1
            if alarmstatuscount[2] == factory and alarmstatuscount[3] == NOTPROCESSED and alarmstatuscount[0] == STOPPAGE:
                undisposed += 1
                alarmcount += 1
        Smartandenvironmentallyfriendlysheet0[stoppagealarmC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[stoppagedisposedD] = disposed
        Smartandenvironmentallyfriendlysheet0[stoppageundisposedE] = undisposed

    # 这里做输出excel文件测试，需要注释掉
    # Smartandenvironmentallyfriendly_input_workbook.save(SmartandenvironmentallyfriendlyOUTPUTFILEPATH)
    # 分厂工况类报警详情处理
    if (cems_operatecondition != 0):
        cems_operatecondition_string = cems_instance.operatecondition(cemsarray, factory)
        # 将返回的字符串输入到日报表对应位置
        Smartandenvironmentallyfriendlysheet0[operateconditionfactoryA] = factory
        Smartandenvironmentallyfriendlysheet0[operateconditionalarmdetailF] = cems_operatecondition_string
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for alarmstatuscount in cemsarray:
            if alarmstatuscount[2] == factory and (alarmstatuscount[3] == PROCESSED or alarmstatuscount[3] == APPROVED) and alarmstatuscount[
                0] == OPERATECONDITION:
                disposed += 1
                alarmcount += 1
            if alarmstatuscount[2] == factory and alarmstatuscount[3] == NOTPROCESSED and alarmstatuscount[
                0] == OPERATECONDITION:
                undisposed += 1
                alarmcount += 1
        Smartandenvironmentallyfriendlysheet0[operateconditionalarmC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[operateconditiondisposedD] = disposed
        Smartandenvironmentallyfriendlysheet0[operateconditionundisposedE] = undisposed
    # 这里做输出excel文件测试，需要注释掉
    # Smartandenvironmentallyfriendly_input_workbook.save(SmartandenvironmentallyfriendlyOUTPUTFILEPATH)
    if (cems_cemsfault != 0):
        cems_cemsfault_string = cems_instance.cemsfault(cemsarray,factory)
        Smartandenvironmentallyfriendlysheet0[cemsfaultfactoryA] = factory
        Smartandenvironmentallyfriendlysheet0[cemsfaultalarmdetailF] = cems_cemsfault_string
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for alarmstatuscount in cemsarray:
            if alarmstatuscount[2] == factory and (alarmstatuscount[3] == PROCESSED or alarmstatuscount[3] == APPROVED) and alarmstatuscount[
                0] == CEMSFAULT:
                disposed += 1
                alarmcount += 1
            if alarmstatuscount[2] == factory and alarmstatuscount[3] == NOTPROCESSED and alarmstatuscount[
                0] == CEMSFAULT:
                undisposed += 1
                alarmcount += 1
        Smartandenvironmentallyfriendlysheet0[cemsfaultalarmC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[cemsfaultdisposedD] = disposed
        Smartandenvironmentallyfriendlysheet0[cemsfaultundisposedE] = undisposed

def inorgnise_dispose(inorganizationarray,Smartandenvironmentallyfriendlysheet0,factory,gfA,gfC,gfD,gfE,gfF,gfG,gfH,gfI,
                      TSPA,TSPC,TSPD,TSPE,TSPF,
                      TVOCA,TVOCC,TVOCD,TVOCE,TVOCF):
    inorgnise_gf = 0
    inorgnise_TSP = 0
    inorgnise_TVOC = 0
    for row in inorganizationarray:
        if row[1] == GOVERNANCEFACILITIES and row[2] == factory:
            inorgnise_gf = inorgnise_gf + 1
        elif row[1] == "TSP" and row[2] == factory:
            inorgnise_TSP = inorgnise_TSP + 1
        elif row[1] == TVOC and row[2] == factory:
            inorgnise_TVOC = inorgnise_TVOC + 1
    inorganization_instance = inorganization.Inorganization()
    if inorgnise_gf != 0:
        inorgnise_gf_string = inorganization_instance.governance(inorganizationarray,factory)
        # 将返回的字符串输入到日报表对应位置
        Smartandenvironmentallyfriendlysheet0[gfA] = factory
        Smartandenvironmentallyfriendlysheet0[gfI] = inorgnise_gf_string
        unlinkage = 0
        differentialpressureoverrun = 0
        fanscurrentoverrun = 0
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for alarmtypecount in inorganizationarray:
            if alarmtypecount[3] == UNLINKAGE and alarmtypecount[2] == factory:
                unlinkage += 1
                alarmcount += 1
            if alarmtypecount[3] == DIFFERENTIALPRESSUREOVERRUN and alarmtypecount[2] == factory:
                differentialpressureoverrun += 1
                alarmcount += 1
            if alarmtypecount[3] == FANSCURRENTOVERRUN and alarmtypecount[2] == factory:
                fanscurrentoverrun += 1
                alarmcount += 1
            if alarmtypecount[2] == factory and alarmtypecount[1] == GOVERNANCEFACILITIES:
                if alarmtypecount[4] == PROCESSED or alarmtypecount[4] == APPROVED:
                    disposed += 1
                else:
                    undisposed += 1
        Smartandenvironmentallyfriendlysheet0[gfC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[gfF] = unlinkage
        Smartandenvironmentallyfriendlysheet0[gfH] = differentialpressureoverrun
        Smartandenvironmentallyfriendlysheet0[gfG] = fanscurrentoverrun
        Smartandenvironmentallyfriendlysheet0[gfD] = disposed
        Smartandenvironmentallyfriendlysheet0[gfE] = undisposed

    if inorgnise_TSP != 0:
        inorgnise_TSP_string = inorganization_instance.TSP(inorganizationarray,factory)
        Smartandenvironmentallyfriendlysheet0[TSPA] = factory
        Smartandenvironmentallyfriendlysheet0[TSPF] = inorgnise_TSP_string
        disposed = 0
        undisposed = 0
        alarmcount = 0
        for count in inorganizationarray:
            if count[2] == factory and count[1] == "TSP":
                alarmcount += 1
                if count[4] == PROCESSED or count[4] == APPROVED and count[2] == factory:
                    disposed += 1
                else:
                    undisposed +=1
        Smartandenvironmentallyfriendlysheet0[TSPC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[TSPD] = disposed
        Smartandenvironmentallyfriendlysheet0[TSPE] = undisposed

    if inorgnise_TVOC != 0 :
        inorgnise_TVOC_string = inorganization_instance.TVOC(inorganizationarray,factory)
        Smartandenvironmentallyfriendlysheet0[TVOCA] = factory
        Smartandenvironmentallyfriendlysheet0[TVOCF] = inorgnise_TVOC_string

        disposed = 0
        undisposed = 0
        alarmcount = 0
        for count in inorganizationarray:
            if count[2] == factory and count[1] == TVOC:
                alarmcount += 1
                if count[4] == PROCESSED or count[4] == APPROVED and count[2] == factory:
                    disposed += 1
                else:
                    undisposed += 1
        Smartandenvironmentallyfriendlysheet0[TVOCC] = alarmcount
        Smartandenvironmentallyfriendlysheet0[TVOCD] = disposed
        Smartandenvironmentallyfriendlysheet0[TVOCE] = undisposed


def main():
    # 定义cems的二维数组，一共6列，从0-5分别包含内容的标题为：类型、设备名称、所属分厂、报警状态、报警详情、报警原因
    cemsarray = np.array([[EVENTTYPE,EQUIPMENTNAME,FACTORY,ALARMSTATUS,ALARMDETAILS,ALARMCAUSE]])

    # 武钢系统表读入
    onlinesheet0 = pd.read_excel(ONLINEFILEPATH, header=None)

    # 用openpyxl进行带格式的读入写出
    Smartandenvironmentallyfriendly_input_workbook = load_workbook(SmartandenvironmentallyfriendlyINPUTFILEPATH)
    Smartandenvironmentallyfriendlysheet0 = Smartandenvironmentallyfriendly_input_workbook["智慧生态环保系统运行日报"]

    first_row = onlinesheet0.iloc[0]
    # <editor-fold desc='寻找导入表各列位置变量'>
    columncout = 0
    eventtypecolumn = 0
    equipmentnamecolumn = 0
    equipmenttypecolumn = 0
    factorycolumn = 0
    productionprocesscolumn = 0
    alarmtypecolumn = 0
    alarmstatuscolumn = 0
    alarmdetailscolumn = 0
    starttimecolumn = 0
    alarmcausecolumn = 0
    # </editor-fold>

    for elemt in first_row:
        if(elemt == EVENTTYPE):
            eventtypecolumn = columncout
        if(elemt == EQUIPMENTNAME):
            equipmentnamecolumn = columncout
        if(elemt == EQUIPMENTTYPE):
            equipmenttypecolumn = columncout
        if (elemt == FACTORY):
            factorycolumn = columncout
        if (elemt == PRODUCTIONPROCESS):
            productionprocesscolumn = columncout
        if (elemt == ALARMTYPE):
            alarmtypecolumn = columncout
        if (elemt == ALARMSTATUS):
            alarmstatuscolumn = columncout
        if (elemt == ALARMDETAILS):
            alarmdetailscolumn = columncout
        if (elemt == STARTTIME):
            starttimecolumn = columncout
        if (elemt == ALARMCAUSE):
            alarmcausecolumn = columncout
        columncout = columncout+1

    # 将对应值输入到cemsarray中
    for row in onlinesheet0.iloc[1:].itertuples(index=False):
        if(row[equipmenttypecolumn] == "CEMS"):
            newrow = [row[eventtypecolumn],row[equipmentnamecolumn],row[factorycolumn],row[alarmstatuscolumn],row[alarmdetailscolumn],row[alarmcausecolumn]]
            cemsarray = np.vstack((cemsarray,newrow))

    # <editor-fold desc='各厂部cems报警数计数变量'>
    cems_ironmake = 0
    cems_steelmake = 0
    cems_strips = 0
    cems_hotroll = 0
    cems_coldroll = 0
    cems_siliconsteel = 0
    cems_steelelectricity = 0
    cems_transport = 0
    cems_eaep = 0
    cems_winsteel = 0
    # </editor-fold>

    for factory in cemsarray:
        if(factory[2] == IRONMAKE):
            cems_ironmake = cems_ironmake + 1
        elif(factory[2] == STEELMAKE):
            cems_steelmake = cems_steelmake +1
        elif(factory[2] == STRIPS):
            cems_strips = cems_strips + 1
        elif(factory[2] == HOTROLL):
            cems_hotroll = cems_hotroll + 1
        elif(factory[2] == COLDROLL):
            cems_coldroll = cems_coldroll + 1
        elif(factory[2] == SILICONSTEEL):
            cems_siliconsteel = cems_siliconsteel + 1
        elif(factory[2] == STEELELECTRICITY):
            cems_steelelectricity = cems_steelelectricity + 1
        elif(factory[2] == TRANSPORT):
            cems_transport = cems_transport + 1
        elif(factory[2] == EAEP):
            cems_eaep = cems_eaep + 1
        elif(factory[2] == WINSTEEL):
            cems_winsteel = cems_winsteel + 1

    # <editor-fold desc='调用cems_dispose函数将不同厂部内容填入模板表'>
    if (cems_ironmake != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,IRONMAKE,"A7","C7","D7","E7","F7","A20","C20","D20","E20","F20","A33","C33","D33","E33","F33","A46","C46","D46","E46","F46")
    if (cems_steelmake != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,STEELMAKE,"A8","C8","D8","E8","F8","A21","C21","D21","E21","F21","A34","C34","D34","E34","F34","A47","C47","D47","E47","F47")
    if (cems_strips != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,STRIPS,"A9","C9","D9","E9","F9","A22","C22","D22","E22","F22","A35","C35","D35","E35","F35","A48","C48","D48","E48","F48")
    if (cems_hotroll != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,HOTROLL,"A10","C10","D10","E10","F10","A23","C23","D23","E23","F23","A36","C36","D36","E36","F36","A49","C49","D49","E49","F49")
    if (cems_coldroll != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,COLDROLL,"A11","C11","D11","E11","F11","A24","C24","D24","E24","F24","A37","C37","D37","E37","F37","A50","C50","D50","E50","F50")
    if (cems_siliconsteel != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,SILICONSTEEL,"A12","C12","D12","E12","F12","A25","C25","D25","E25","F25","A38","C38","D38","E38","F38","A51","C51","D51","E51","F51")
    if (cems_steelelectricity != 0):
        cems_dispose(cemsarray, Smartandenvironmentallyfriendlysheet0, STEELELECTRICITY, "A13", "C13", "D13", "E13", "F13", "A26", "C26", "D26", "E26", "F26", "A39", "C39", "D39", "E39", "F39","A52","C52","D52","E52","F52")
    if (cems_transport != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,TRANSPORT,"A14","C14","D14","E14","F14","A27","C27","D27","E27","F27","A40","C40","D40","E40","F40","A53","C53","D53","E53","F53")
    if (cems_eaep != 0):
        cems_dispose(cemsarray,Smartandenvironmentallyfriendlysheet0,EAEP,"A15","C15","D15","E15","F15","A28","C28","D28","E28","F28","A41","C41","D41","E41","F41","A54","C54","D54","E54","F54")
    if (cems_winsteel != 0):
        cems_dispose(cemsarray, Smartandenvironmentallyfriendlysheet0, WINSTEEL, "A16", "C16", "D16", "E16", "F16", "A29","C29", "D29", "E29", "F29", "A42", "C42", "D42", "E42", "F42", "A55", "C55", "D55", "E55", "F55")

    # </editor-fold>

    # ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # 上面代码完成cems部分的输入
    # 下面开始处理无组织监测监控报警及处理情况（环保设备、TSP、TVOC）
    # 先将无组织数组读入，信息读全，再将无组织数组按照报警类型分为“未联动开启”数组、“风机电流超限”数组、“压差超限”数组
    inorganizationarray = [[EQUIPMENTNAME,EQUIPMENTTYPE,FACTORY,ALARMTYPE,ALARMSTATUS,ALARMDETAILS,STARTTIME,ALARMCAUSE]]
    # <editor-fold desc='寻找无组织表各列名称位置'>
    equipmentnamecolumn = 0
    equipmenttypecolumn = 0
    factorycolumn = 0
    alarmtypecolumn = 0
    alarmstatuscolumn = 0
    alarmdetailscolumn = 0
    starttimecolumn = 0
    alarmcausecolumn = 0
    # </editor-fold>

    columncout = 0
    for count in first_row:
        if count == EQUIPMENTNAME:
            equipmentnamecolumn = columncout
        if count == EQUIPMENTTYPE:
            equipmenttypecolumn = columncout
        if count == FACTORY:
            factorycolumn = columncout
        if count == ALARMTYPE:
            alarmtypecolumn = columncout
        if count == ALARMSTATUS:
            alarmstatuscolumn = columncout
        if count == ALARMDETAILS:
            alarmdetailscolumn = columncout
        if count == STARTTIME:
            starttimecolumn = columncout
        if count == ALARMCAUSE:
            alarmcausecolumn = columncout
        columncout = columncout + 1

    columncout = 1
    for row in onlinesheet0.iloc[1:].itertuples(index=False):
        if row[equipmenttypecolumn] == GOVERNANCEFACILITIES or row[equipmenttypecolumn] == "TSP" or row[equipmenttypecolumn] == TVOC:
            if pd.isnull(onlinesheet0.at[columncout,alarmcausecolumn]):
                newrow = [row[equipmentnamecolumn],row[equipmenttypecolumn],row[factorycolumn],row[alarmtypecolumn],row[alarmstatuscolumn],row[alarmdetailscolumn],row[starttimecolumn],"待反馈"]
            else:
                newrow = [row[equipmentnamecolumn],row[equipmenttypecolumn],row[factorycolumn],row[alarmtypecolumn],row[alarmstatuscolumn],row[alarmdetailscolumn],row[starttimecolumn],row[alarmcausecolumn]]
            inorganizationarray = np.vstack((inorganizationarray,newrow))
        columncout = columncout + 1

    # <editor-fold desc='无组织计数变量'>
    inorganise_ironmake = 0
    inorganise_steelmake = 0
    inorganise_strips = 0
    inorganise_hotroll = 0
    inorganise_coldroll = 0
    inorganise_siliconsteel = 0
    inorganise_steelelectricity = 0
    inorganise_transport = 0
    inorganise_eaep = 0
    # </editor-fold>

    for row in inorganizationarray:
        if row[2] == IRONMAKE:
            inorganise_ironmake = inorganise_ironmake + 1
        if row[2] == STEELMAKE:
            inorganise_steelmake = inorganise_steelmake + 1
        if row[2] == STRIPS:
            inorganise_strips = inorganise_strips + 1
        if row[2] == HOTROLL:
            inorganise_hotroll = inorganise_hotroll + 1
        if row[2] == COLDROLL:
            inorganise_coldroll = inorganise_coldroll + 1
        if row[2] == SILICONSTEEL:
            inorganise_siliconsteel = inorganise_siliconsteel + 1
        if row[2] == STEELELECTRICITY:
            inorganise_steelelectricity = inorganise_steelelectricity + 1
        if row[2] == TRANSPORT:
            inorganise_transport = inorganise_transport + 1
        if row[2] == EAEP:
            inorganise_eaep = inorganise_eaep + 1

    # <editor-fold desc='调用inorgnise_dispose方法处理各厂无组织信息并填入模板表内'>
    if inorganise_ironmake != 0:
        inorgnise_dispose(inorganizationarray,Smartandenvironmentallyfriendlysheet0,IRONMAKE,"A62","C62","D62","E62","F62","G62","H62","I62","A74","C74","D74","E74","F74","A86","C86","D86","E86","F86")
    if inorganise_steelmake != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, STEELMAKE, "A63", "C63", "D63","E63", "F63", "G63", "H63", "I63", "A75", "C75", "D75", "E75", "F75", "A87", "C87", "D87","E87", "F87")
    if inorganise_strips != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, STRIPS, "A64", "C64", "D64","E64", "F64", "G64", "H64", "I64", "A76", "C76", "D76", "E76", "F76", "A88", "C88", "D88","E88", "F88")
    if inorganise_hotroll != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, HOTROLL, "A65", "C65", "D65","E65", "F65", "G65", "H65", "I65", "A77", "C77", "D77", "E77", "F77", "A89", "C89", "D89","E89", "F89")
    if inorganise_coldroll != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, COLDROLL, "A66", "C66", "D66","E66", "F66", "G66", "H66", "I66", "A78", "C78", "D78", "E78", "F78", "A90", "C90", "D90","E90", "F90")
    if inorganise_siliconsteel != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, SILICONSTEEL, "A67", "C67", "D67","E67", "F67", "G67", "H67", "I67", "A79", "C79", "D79", "E79", "F79", "A91", "C91", "D91","E91", "F91")
    if inorganise_steelelectricity != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0,STEELELECTRICITY , "A68", "C68", "D68","E68", "F68", "G68", "H68", "I68", "A80", "C80", "D80", "E80", "F80", "A92", "C92", "D92","E92", "F92")
    if inorganise_transport != 0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, TRANSPORT, "A69", "C69", "D69","E69", "F69", "G69", "H69", "I69", "A81", "C81", "D81", "E81", "F81", "A93", "C93", "D93","E93", "F93")
    if inorganise_eaep !=0:
        inorgnise_dispose(inorganizationarray, Smartandenvironmentallyfriendlysheet0, EAEP, "A70", "C70", "D70","E70", "F70", "G70", "H70", "I70", "A82", "C82", "D82", "E82", "F82", "A94", "C94", "D94","E94", "F94")
    # </editor-fold>

    # <editor-fold desc='CEMS各种数量统计填入汇总单元格'>
    allcemsstring = "1.报警总数"
    allcemsalarm = 0
    allcalibration = 0
    # <editor-fold desc='检校类型各个厂部计数变量定义'>
    caim = 0
    casm = 0
    cast = 0
    cahr = 0
    cacr = 0
    cass = 0
    case = 0
    catr = 0
    caea = 0
    cawi = 0
    # </editor-fold>
    allstoppage = 0
    # <editor-fold desc='启停类型各个厂部计数变量定义'>
    stim = 0
    stsm = 0
    stst = 0
    sthr = 0
    stcr = 0
    stss = 0
    stse = 0
    sttr = 0
    stea = 0
    stwi = 0
    # </editor-fold>
    alloperatecondition = 0
    # <editor-fold desc='工况类型各个厂部计数变量定义'>
    opim = 0
    opsm = 0
    opst = 0
    ophr = 0
    opcr = 0
    opss = 0
    opse = 0
    optr = 0
    opea = 0
    opwi = 0
    # </editor-fold>
    allcemsfault = 0
    # <editor-fold desc='CEMS故障类型各厂部计数变量定义'>
    cfim = 0
    cfsm = 0
    cfst = 0
    cfhr = 0
    cfcr = 0
    cfss = 0
    cfse = 0
    cftr = 0
    cfea = 0
    cfwi = 0
    # </editor-fold>
    for row in cemsarray[1:]:
        allcemsalarm += 1
        if row[0] == CALIBRATION:
            allcalibration += 1
            if row[2] == IRONMAKE:
                caim += 1
            elif row[2] == STEELMAKE:
                casm += 1
            elif row[2] == STRIPS:
                cast += 1
            elif row[2] == HOTROLL:
                cahr += 1
            elif row[2] == COLDROLL:
                cacr += 1
            elif row[2] == SILICONSTEEL:
                cass += 1
            elif row[2] == STEELELECTRICITY:
                case += 1
            elif row[2] == TRANSPORT:
                catr += 1
            elif row[2] == EAEP:
                caea += 1
            elif row[2] == WINSTEEL:
                cawi += 1
        if row[0] == STOPPAGE:
            allstoppage += 1
            if row[2] == IRONMAKE:
                stim += 1
            elif row[2] == STEELMAKE:
                stsm += 1
            elif row[2] == STRIPS:
                stst += 1
            elif row[2] == HOTROLL:
                sthr += 1
            elif row[2] == COLDROLL:
                stcr += 1
            elif row[2] == SILICONSTEEL:
                stss += 1
            elif row[2] == STEELELECTRICITY:
                stse += 1
            elif row[2] == TRANSPORT:
                sttr += 1
            elif row[2] == EAEP:
                stea += 1
            elif row[2] == WINSTEEL:
                stwi += 1
        if row[0] == OPERATECONDITION:
            alloperatecondition += 1
            if row[2] == IRONMAKE:
                opim+= 1
            elif row[2] == STEELMAKE:
                opsm += 1
            elif row[2] == STRIPS:
                opst += 1
            elif row[2] == HOTROLL:
                ophr += 1
            elif row[2] == COLDROLL:
                opcr += 1
            elif row[2] == SILICONSTEEL:
                opss += 1
            elif row[2] == STEELELECTRICITY:
                opse += 1
            elif row[2] == TRANSPORT:
                optr += 1
            elif row[2] == EAEP:
                opea += 1
            elif row[2] == WINSTEEL:
                opwi += 1
        if row[0] == CEMSFAULT:
            allcemsfault += 1
            if row[2] == IRONMAKE:
                cfim+= 1
            elif row[2] == STEELMAKE:
                cfsm += 1
            elif row[2] == STRIPS:
                cfst += 1
            elif row[2] == HOTROLL:
                cfhr += 1
            elif row[2] == COLDROLL:
                cfcr += 1
            elif row[2] == SILICONSTEEL:
                cfss += 1
            elif row[2] == STEELELECTRICITY:
                cfse += 1
            elif row[2] == TRANSPORT:
                cftr += 1
            elif row[2] == EAEP:
                cfea += 1
            elif row[2] == WINSTEEL:
                cfwi += 1
    allcemsstring += str(allcemsalarm) + "个，其中"
    if allcalibration != 0:
        allcemsstring += "检校报警" + str(allcalibration) + "个（"
        if caim != 0:
            allcemsstring += IRONMAKE + str(caim) + "个、"
        if casm != 0:
            allcemsstring += STEELMAKE + str(casm) + "个、"
        if cast != 0:
            allcemsstring += STRIPS + str(cast) + "个、"
        if cahr != 0:
            allcemsstring += HOTROLL + str(cahr) + "个、"
        if cacr != 0:
            allcemsstring += COLDROLL + str(cacr) + "个、"
        if cass != 0:
            allcemsstring += SILICONSTEEL + str(cass) + "个、"
        if case != 0:
            allcemsstring += STEELELECTRICITY + str(case) + "个、"
        if catr != 0:
            allcemsstring += TRANSPORT + str(catr) + "个、"
        if caea != 0:
            allcemsstring += EAEP + str(caea) + "个、"
        if cawi != 0:
            allcemsstring += WINSTEEL + str(cawi) + "个、"
        allcemsstring = allcemsstring.rstrip(allcemsstring[-1])
        allcemsstring += "）"
    if allstoppage != 0:
        allcemsstring += "，启停炉（机）报警" + str(allstoppage) + "个（"
        if stim != 0:
            allcemsstring += IRONMAKE + str(stim) + "个、"
        if stsm != 0:
            allcemsstring += STEELMAKE + str(stsm) + "个、"
        if stst != 0:
            allcemsstring += STRIPS + str(stst) + "个、"
        if sthr != 0:
            allcemsstring += HOTROLL + str(sthr) + "个、"
        if stcr != 0:
            allcemsstring += COLDROLL + str(stcr) + "个、"
        if stss != 0:
            allcemsstring += SILICONSTEEL + str(stss) + "个、"
        if stse != 0:
            allcemsstring += STEELELECTRICITY + str(stse) + "个、"
        if sttr != 0:
            allcemsstring += TRANSPORT + str(sttr) + "个、"
        if stea != 0:
            allcemsstring += EAEP + str(stea) + "个、"
        if stwi != 0:
            allcemsstring += WINSTEEL + str(stwi) + "个、"
        allcemsstring = allcemsstring.rstrip(allcemsstring[-1])
        allcemsstring += "）"
    if alloperatecondition != 0:
        allcemsstring += "，工况异常报警" + str(alloperatecondition) + "个（"
        if opim != 0:
            allcemsstring += IRONMAKE + str(opim) + "个、"
        if opsm != 0:
            allcemsstring += STEELMAKE + str(opsm) + "个、"
        if opst != 0:
            allcemsstring += STRIPS + str(opst) + "个、"
        if ophr != 0:
            allcemsstring += HOTROLL + str(ophr) + "个、"
        if opcr != 0:
            allcemsstring += COLDROLL + str(opcr) + "个、"
        if opss != 0:
            allcemsstring += SILICONSTEEL + str(opss) + "个、"
        if opse != 0:
            allcemsstring += STEELELECTRICITY + str(opse) + "个、"
        if optr != 0:
            allcemsstring += TRANSPORT + str(optr) + "个、"
        if opea != 0:
            allcemsstring += EAEP + str(opea) + "个、"
        if opwi != 0:
            allcemsstring += WINSTEEL + str(opwi) + "个、"
        allcemsstring = allcemsstring.rstrip(allcemsstring[-1])
        allcemsstring += "）"
    if allcemsfault != 0:
        allcemsstring += "，在线设备异常报警" + str(allcemsfault) + "个（"
        if cfim != 0:
            allcemsstring += IRONMAKE + str(cfim) + "个、"
        if cfsm != 0:
            allcemsstring += STEELMAKE + str(cfsm) + "个、"
        if cfst != 0:
            allcemsstring += STRIPS + str(cfst) + "个、"
        if cfhr != 0:
            allcemsstring += HOTROLL + str(cfhr) + "个、"
        if cfcr != 0:
            allcemsstring += COLDROLL + str(cfcr) + "个、"
        if cfss != 0:
            allcemsstring += SILICONSTEEL + str(cfss) + "个、"
        if cfse != 0:
            allcemsstring += STEELELECTRICITY + str(cfse) + "个、"
        if cftr != 0:
            allcemsstring += TRANSPORT + str(cftr) + "个、"
        if cfea != 0:
            allcemsstring += EAEP + str(cfea) + "个、"
        if cfwi != 0:
            allcemsstring += WINSTEEL + str(cfwi) + "个、"
        allcemsstring = allcemsstring.rstrip(allcemsstring[-1])
        allcemsstring += "）"
    allcemsstring += "\n" + "2."
    # 获取当前日期计算出明天日期
    tomorrow = datetime.now() + timedelta(days=1)
    tomorrowstring = tomorrow.strftime("%m月%d日")
    if tomorrowstring[0] == "0":
        tomorrowstring = tomorrowstring[1:]
    allcemsstring += tomorrowstring + "计划校准"
    # 读校准计划并将内容加入allcemsstring
    calibrationplanworkbook = load_workbook(calibrationplanFILEPATH)
    calibrationplanworkbooksheet0 = calibrationplanworkbook["汇总"]
    # <editor-fold desc='判断明天是周几并找到对应单元格位置'>
    planrow = 0
    if tomorrow.weekday() == 0:
        planrow = 1
    elif tomorrow.weekday() == 1:
        planrow = 2
    elif tomorrow.weekday() == 2:
        planrow = 3
    elif tomorrow.weekday() == 3:
        planrow = 4
    elif tomorrow.weekday() == 4:
        planrow = 5
    elif tomorrow.weekday() == 5:
        planrow = 6
    elif tomorrow.weekday() == 6:
        planrow = 7
    # </editor-fold>
    allcemsstring += calibrationplanworkbooksheet0.cell(row = planrow+1,column = 6).value
    # 写入
    Smartandenvironmentallyfriendlysheet0["A4"] = allcemsstring
    #</editor-fold>

    # <editor-fold desc='无组织各种数量统计填入汇总单元格'>
    allorgnisestring = "1.报警总数"
    allorgnisealarm = 0
    allgovernequipment = 0
    # <editor-fold desc='环保设备各厂报警计数变量'>
    geim = 0
    gesm = 0
    gest = 0
    gehr = 0
    gecr = 0
    gess = 0
    gese = 0
    getr = 0
    geea = 0
    # </editor-fold>
    allTSP = 0
    # <editor-fold desc='TSP各厂报警计数变量'>
    TSPim = 0
    TSPsm = 0
    TSPst = 0
    TSPhr = 0
    TSPcr = 0
    TSPss = 0
    TSPse = 0
    TSPtr = 0
    TSPea = 0
    # </editor-fold>
    allTVOC = 0
    # <editor-fold desc='TVOC各厂报警计数变量'>
    TVOCim = 0
    TVOCsm = 0
    TVOCst = 0
    TVOChr = 0
    TVOCcr = 0
    TVOCss = 0
    TVOCse = 0
    TVOCtr = 0
    TVOCea = 0
    # </editor-fold>
    for row in inorganizationarray[1:]:
        allorgnisealarm += 1
        if row[1] == GOVERNANCEFACILITIES:
            allgovernequipment += 1
            if row[2] == IRONMAKE:
                geim += 1
            elif row[2] == STEELMAKE:
                gesm += 1
            elif row[2] == STRIPS:
                gest += 1
            elif row[2] == HOTROLL:
                gehr += 1
            elif row[2] == COLDROLL:
                gecr += 1
            elif row[2] == SILICONSTEEL:
                gess += 1
            elif row[2] == STEELELECTRICITY:
                gese += 1
            elif row[2] == TRANSPORT:
                getr += 1
            elif row[2] == EAEP:
                geea += 1
        if row[1] == "TSP":
            allTSP += 1
            if row[2] == IRONMAKE:
                TSPim += 1
            elif row[2] == STEELMAKE:
                TSPsm += 1
            elif row[2] == STRIPS:
                TSPst += 1
            elif row[2] == HOTROLL:
                TSPhr += 1
            elif row[2] == COLDROLL:
                TSPcr += 1
            elif row[2] == SILICONSTEEL:
                TSPss += 1
            elif row[2] == STEELELECTRICITY:
                TSPse += 1
            elif row[2] == TRANSPORT:
                TSPtr += 1
            elif row[2] == EAEP:
                TSPea += 1
        if row[1] == TVOC:
            allTVOC += 1
            if row[2] == IRONMAKE:
                TVOCim += 1
            elif row[2] == STEELMAKE:
                TVOCsm += 1
            elif row[2] == STRIPS:
                TVOCst += 1
            elif row[2] == HOTROLL:
                TVOChr += 1
            elif row[2] == COLDROLL:
                TVOCcr += 1
            elif row[2] == SILICONSTEEL:
                TVOCss += 1
            elif row[2] == STEELELECTRICITY:
                TVOCse += 1
            elif row[2] == TRANSPORT:
                TVOCtr += 1
            elif row[2] == EAEP:
                TVOCea += 1
    allorgnisestring += str(allorgnisealarm) + "个，其中"
    if allgovernequipment != 0:
        allorgnisestring += "环保设备异常报警" + str(allgovernequipment) + "个（"
        if geim != 0:
            allorgnisestring += IRONMAKE + str(geim) + "个、"
        if gesm != 0:
            allorgnisestring += STEELMAKE + str(gesm) + "个、"
        if gest != 0:
            allorgnisestring += STRIPS + str(gest) + "个、"
        if gehr != 0:
            allorgnisestring += HOTROLL + str(gehr) + "个、"
        if gecr != 0:
            allorgnisestring += COLDROLL + str(gecr) + "个、"
        if gess != 0:
            allorgnisestring += SILICONSTEEL + str(gess) + "个、"
        if gese != 0:
            allorgnisestring += STEELELECTRICITY + str(gese) + "个、"
        if getr != 0:
            allorgnisestring += TRANSPORT + str(getr) + "个、"
        if geea != 0:
            allorgnisestring += EAEP + str(geea) + "个、"
        allorgnisestring = allorgnisestring.rstrip(allorgnisestring[-1])
        allorgnisestring += "），"
    if allTSP != 0:
        allorgnisestring += "TSP异常报警" + str(allTSP) + "个（"
        if TSPim != 0:
            allorgnisestring += IRONMAKE + str(TSPim) + "个、"
        if TSPsm != 0:
            allorgnisestring += STEELMAKE + str(TSPsm) + "个、"
        if TSPst != 0:
            allorgnisestring += STRIPS + str(TSPst) + "个、"
        if TSPhr != 0:
            allorgnisestring += HOTROLL + str(TSPhr) + "个、"
        if TSPcr != 0:
            allorgnisestring += COLDROLL + str(TSPcr) + "个、"
        if TSPss != 0:
            allorgnisestring += SILICONSTEEL + str(TSPss) + "个、"
        if TSPse != 0:
            allorgnisestring += STEELELECTRICITY + str(TSPse) + "个、"
        if TSPtr != 0:
            allorgnisestring += TRANSPORT + str(TSPtr) + "个、"
        if TSPea != 0:
            allorgnisestring += EAEP + str(TSPea) + "个、"
        allorgnisestring = allorgnisestring.rstrip(allorgnisestring[-1])
        allorgnisestring += "），"
    if allTVOC != 0:
        allorgnisestring += "TVOC异常报警" + str(allTVOC) + "个（"
        if TVOCim != 0:
            allorgnisestring += IRONMAKE + str(TVOCim) + "个、"
        if TVOCsm != 0:
            allorgnisestring += STEELMAKE + str(TVOCsm) + "个、"
        if TVOCst != 0:
            allorgnisestring += STRIPS + str(TVOCst) + "个、"
        if TVOChr != 0:
            allorgnisestring += HOTROLL + str(TVOChr) + "个、"
        if TVOCcr != 0:
            allorgnisestring += COLDROLL + str(TVOCcr) + "个、"
        if TVOCss != 0:
            allorgnisestring += SILICONSTEEL + str(TVOCss) + "个、"
        if TVOCse != 0:
            allorgnisestring += STEELELECTRICITY + str(TVOCse) + "个、"
        if TVOCtr != 0:
            allorgnisestring += TRANSPORT + str(TVOCtr) + "个、"
        if TVOCea != 0:
            allorgnisestring += EAEP + str(TVOCea) + "个、"
        allorgnisestring = allorgnisestring.rstrip(allorgnisestring[-1])
        allorgnisestring += "），"
    allorgnisestring = allorgnisestring.rstrip(allorgnisestring[-1])
    allorgnisestring += "\n" + "2.设备离线情况：空气微站离线3个（炼铁厂3个）"
    Smartandenvironmentallyfriendlysheet0['A58'] = allorgnisestring
    # </editor-fold>

    # print(cemsarray)
    ppt_summary(cemsarray,Smartandenvironmentallyfriendlysheet0,"A111","A112")

    # delete_empty_rows(Smartandenvironmentallyfriendlysheet0)
    # xlsx文件输出测试
    Smartandenvironmentallyfriendly_input_workbook.save(SmartandenvironmentallyfriendlyOUTPUTFILEPATH)

if __name__ == "__main__":
    main()
